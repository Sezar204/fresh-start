import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

class DataImporter:
    @staticmethod
    def generate_template(entity: str) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{entity.title()} Import"

        headers_map = {
            "production-lines": ["Code*", "Name*", "Type (discrete/process/batch)", "Capacity Per Hour*", "Unit", "Changeover Minutes"],
            "machines": ["Code*", "Name*", "Line Code", "Criticality (low/medium/high/critical)", "Status", "Capacity", "Unit"],
            "products": ["SKU*", "Name*", "Category", "Unit of Measure*", "Standard Cost*", "Selling Price*", "Type (finished/semi-finished)"],
            "raw-materials": ["Code*", "Name*", "Category", "Unit of Measure*", "Standard Cost*", "Safety Stock*", "Reorder Point*", "Lead Time Days*"],
            "suppliers": ["Code*", "Name*", "Contact Name", "Contact Email", "Contact Phone", "Payment Terms Days", "Rating (1-5)"],
            "customers": ["Code*", "Name*", "Type (b2b/b2c/distributor)", "Priority (1-5)", "Credit Limit", "Payment Terms Days"],
            "warehouses": ["Code*", "Name*", "Type (raw_material/wip/finished_goods/general)", "Total Capacity", "Capacity Unit"],
        }

        headers = headers_map.get(entity, ["Code*", "Name*", "Notes"])

        # Row 1: Headers (Bold, White text, Dark Blue background)
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def parse_and_validate(entity: str, content: bytes) -> dict:
        try:
            df = pd.read_excel(io.BytesIO(content))
            df = df.fillna("")
            records = df.to_dict(orient="records")
            return {
                "total": len(records),
                "valid": len(records),
                "errors": [],
                "preview": records[:10]
            }
        except Exception as e:
            return {
                "total": 0,
                "valid": 0,
                "errors": [{"field": "file", "message": f"Excel parsing error: {str(e)}"}],
                "preview": []
            }
